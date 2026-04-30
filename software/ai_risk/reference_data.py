from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Iterable
from urllib.request import Request, urlopen
import xml.etree.ElementTree as ET

from openpyxl import load_workbook

from ai_risk.utils.io import write_rows, write_text

CSF_DOWNLOAD_URL = "https://csrc.nist.gov/extensions/nudp/services/json/csf/download?olirids=all"

ISO_DOMAINS: dict[str, tuple[str, int]] = {
    "A.5": ("Organizational", 37),
    "A.6": ("People", 8),
    "A.7": ("Physical", 14),
    "A.8": ("Technological", 34),
}

ISO_CONTROL_CATALOG: list[tuple[str, str, str]] = [
    ("A.5.1", "Policies for Information Security", "Organizational"),
    ("A.5.2", "Information Security Roles and Responsibilities", "Organizational"),
    ("A.5.3", "Segregation of Duties", "Organizational"),
    ("A.5.4", "Management Responsibilities", "Organizational"),
    ("A.5.5", "Contact With Authorities", "Organizational"),
    ("A.5.6", "Contact With Special Interest Groups", "Organizational"),
    ("A.5.7", "Threat Intelligence", "Organizational"),
    ("A.5.8", "Information Security in Project Management", "Organizational"),
    ("A.5.9", "Inventory of Information and Other Associated Assets", "Organizational"),
    ("A.5.10", "Acceptable Use of Information and Other Associated Assets", "Organizational"),
    ("A.5.11", "Return of Assets", "Organizational"),
    ("A.5.12", "Classification of Information", "Organizational"),
    ("A.5.13", "Labelling of Information", "Organizational"),
    ("A.5.14", "Information Transfer", "Organizational"),
    ("A.5.15", "Access Control", "Organizational"),
    ("A.5.16", "Identity Management", "Organizational"),
    ("A.5.17", "Authentication Information", "Organizational"),
    ("A.5.18", "Access Rights", "Organizational"),
    ("A.5.19", "Information Security in Supplier Relationships", "Organizational"),
    ("A.5.20", "Addressing Information Security Within Supplier Agreements", "Organizational"),
    ("A.5.21", "Managing Information Security in the ICT Supply Chain", "Organizational"),
    ("A.5.22", "Monitoring, Review and Change Management of Supplier Services", "Organizational"),
    ("A.5.23", "Information Security for Use of Cloud Services", "Organizational"),
    ("A.5.24", "Information Security Incident Management Planning and Preparation", "Organizational"),
    ("A.5.25", "Assessment and Decision on Information Security Events", "Organizational"),
    ("A.5.26", "Response to Information Security Incidents", "Organizational"),
    ("A.5.27", "Learning From Information Security Incidents", "Organizational"),
    ("A.5.28", "Collection of Evidence", "Organizational"),
    ("A.5.29", "Information Security During Disruption", "Organizational"),
    ("A.5.30", "ICT Readiness for Business Continuity", "Organizational"),
    ("A.5.31", "Legal, Statutory, Regulatory and Contractual Requirements", "Organizational"),
    ("A.5.32", "Intellectual Property Rights", "Organizational"),
    ("A.5.33", "Protection of Records", "Organizational"),
    ("A.5.34", "Privacy and Protection of PII", "Organizational"),
    ("A.5.35", "Independent Review of Information Security", "Organizational"),
    ("A.5.36", "Compliance With Policies, Rules and Standards for Information Security", "Organizational"),
    ("A.5.37", "Documented Operating Procedures", "Organizational"),
    ("A.6.1", "Screening", "People"),
    ("A.6.2", "Terms and Conditions of Employment", "People"),
    ("A.6.3", "Information Security Awareness, Education and Training", "People"),
    ("A.6.4", "Disciplinary Process", "People"),
    ("A.6.5", "Responsibilities After Termination or Change of Employment", "People"),
    ("A.6.6", "Confidentiality or Non-Disclosure Agreements", "People"),
    ("A.6.7", "Remote Working", "People"),
    ("A.6.8", "Information Security Event Reporting", "People"),
    ("A.7.1", "Physical Security Perimeters", "Physical"),
    ("A.7.2", "Physical Entry", "Physical"),
    ("A.7.3", "Securing Offices, Rooms and Facilities", "Physical"),
    ("A.7.4", "Physical Security Monitoring", "Physical"),
    ("A.7.5", "Protecting Against Physical and Environmental Threats", "Physical"),
    ("A.7.6", "Working In Secure Areas", "Physical"),
    ("A.7.7", "Clear Desk and Clear Screen", "Physical"),
    ("A.7.8", "Equipment Siting and Protection", "Physical"),
    ("A.7.9", "Security of Assets Off-Premises", "Physical"),
    ("A.7.10", "Storage Media", "Physical"),
    ("A.7.11", "Supporting Utilities", "Physical"),
    ("A.7.12", "Cabling Security", "Physical"),
    ("A.7.13", "Equipment Maintenance", "Physical"),
    ("A.7.14", "Secure Disposal or Re-Use of Equipment", "Physical"),
    ("A.8.1", "User Endpoint Devices", "Technological"),
    ("A.8.2", "Privileged Access Rights", "Technological"),
    ("A.8.3", "Information Access Restriction", "Technological"),
    ("A.8.4", "Access to Source Code", "Technological"),
    ("A.8.5", "Secure Authentication", "Technological"),
    ("A.8.6", "Capacity Management", "Technological"),
    ("A.8.7", "Protection Against Malware", "Technological"),
    ("A.8.8", "Management of Technical Vulnerabilities", "Technological"),
    ("A.8.9", "Configuration Management", "Technological"),
    ("A.8.10", "Information Deletion", "Technological"),
    ("A.8.11", "Data Masking", "Technological"),
    ("A.8.12", "Data Leakage Prevention", "Technological"),
    ("A.8.13", "Information Backup", "Technological"),
    ("A.8.14", "Redundancy of Information Processing Facilities", "Technological"),
    ("A.8.15", "Logging", "Technological"),
    ("A.8.16", "Monitoring Activities", "Technological"),
    ("A.8.17", "Clock Synchronization", "Technological"),
    ("A.8.18", "Use of Privileged Utility Programs", "Technological"),
    ("A.8.19", "Installation of Software on Operational Systems", "Technological"),
    ("A.8.20", "Network Security", "Technological"),
    ("A.8.21", "Security of Network Services", "Technological"),
    ("A.8.22", "Segregation of Networks", "Technological"),
    ("A.8.23", "Web Filtering", "Technological"),
    ("A.8.24", "Use of Cryptography", "Technological"),
    ("A.8.25", "Secure Development Life Cycle", "Technological"),
    ("A.8.26", "Application Security Requirements", "Technological"),
    ("A.8.27", "Secure System Architecture and Engineering Principles", "Technological"),
    ("A.8.28", "Secure Coding", "Technological"),
    ("A.8.29", "Security Testing in Development and Acceptance", "Technological"),
    ("A.8.30", "Outsourced Development", "Technological"),
    ("A.8.31", "Separation of Development, Test and Production Environments", "Technological"),
    ("A.8.32", "Change Management", "Technological"),
    ("A.8.33", "Test Information", "Technological"),
    ("A.8.34", "Protection of Information Systems During Audit Testing", "Technological"),
]

METRIC_CATALOG = [
    ("AV_MONETARY", "Normalized asset monetary value"),
    ("EP_FREQ", "Telemetry-derived event frequency"),
    ("EP_EXPLOIT", "CVSS-based exploitability"),
    ("IL_CIA", "Confidentiality/integrity/availability impact"),
    ("HF_AWARENESS", "Human-factor awareness and culture"),
    ("DR", "Detection rate / recall"),
    ("FPR", "False-positive rate"),
    ("AUROC", "Area under ROC curve"),
    ("PR_AUC", "Area under precision-recall curve"),
    ("CAL_BRIER", "Brier score"),
    ("RRL_PORTFOLIO", "Risk-reduction level"),
    ("TCO_3Y", "Three-year total cost of ownership"),
]

UML_CLASSES = [
    "Asset",
    "Threat",
    "Vulnerability",
    "Event",
    "Impact",
    "Risk",
    "RiskTreatment",
    "SecurityRequirement",
    "SecurityObjective",
    "AssetValue",
    "EventPotential",
    "ImpactLevel",
    "RiskReductionLevel",
    "CostMetrics",
    "HumanFactorMetric",
]


@dataclass(frozen=True)
class ReferencePaths:
    iso_catalog: Path
    nist_catalog: Path
    metric_catalog: Path
    mapping_catalog: Path
    uml_schema: Path
    control_catalog: Path
    crosswalk_catalog: Path


def fetch_csf_workbook_bytes(url: str = CSF_DOWNLOAD_URL) -> bytes:
    request = Request(url, headers={"User-Agent": "Mozilla/5.0", "Accept-Encoding": "identity"})
    with urlopen(request, timeout=120) as response:
        return response.read()


def _clean_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\xa0", " ").strip()


def parse_csf_subcategories(workbook_bytes: bytes) -> list[dict[str, str]]:
    wb = load_workbook(filename=BytesIO(workbook_bytes), read_only=True, data_only=True)
    ws = wb["CSF 2.0"]
    subcat_pattern = re.compile(r"^([A-Z]{2}\.[A-Z]{2}-\d{2}):\s*(.+)$")
    current_function = ""
    current_category = ""
    rows: list[dict[str, str]] = []
    for values in ws.iter_rows(min_row=3, values_only=True):
        function_cell, category_cell, subcategory_cell, _, informative_refs = values
        function_text = _clean_cell(function_cell)
        category_text = _clean_cell(category_cell)
        subcategory_text = _clean_cell(subcategory_cell)
        if function_text:
            current_function = function_text
        if category_text:
            current_category = category_text
        if not subcategory_text:
            continue
        match = subcat_pattern.match(subcategory_text)
        if not match:
            continue
        subcat_id, title = match.groups()
        if "[Withdrawn:" in title:
            continue
        linked_iso = sorted(set(re.findall(r"Annex A Controls:\s*([5-8]\.\d+)", _clean_cell(informative_refs))))
        rows.append(
            {
                "id": subcat_id,
                "title": title,
                "function": current_function,
                "category": current_category,
                "informative_references": _clean_cell(informative_refs).replace("\n", " | "),
                "linked_iso_ids": ";".join(f"A.{value}" for value in linked_iso),
            }
        )
    if len(rows) != 106:
        raise ValueError(f"Expected 106 NIST CSF 2.0 subcategories, found {len(rows)}")
    return rows


def generate_iso_controls() -> list[dict[str, str]]:
    rows = [{"id": control_id, "title": title, "domain": domain} for control_id, title, domain in ISO_CONTROL_CATALOG]
    if len(rows) != 93:
        raise ValueError(f"Expected 93 ISO controls, found {len(rows)}")
    return rows


def _maybe_write_rows(path: Path, rows: list[dict[str, object]], fieldnames: list[str], refresh: bool) -> Path:
    if path.exists() and not refresh:
        return path
    return write_rows(path, rows, fieldnames)


def _maybe_write_text(path: Path, content: str, refresh: bool) -> Path:
    if path.exists() and not refresh:
        return path
    return write_text(path, content)


def _read_existing_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def _remove_stale_reference_artifacts(data_dir: Path) -> None:
    for filename in ("tco_params.csv", "usage_baseline.csv", "usage_nsga2.csv"):
        stale_path = data_dir / filename
        if stale_path.exists():
            stale_path.unlink()


def build_metric_catalog() -> list[dict[str, str]]:
    return [{"metric_id": metric_id, "description": description} for metric_id, description in METRIC_CATALOG]


def metric_bundle_for_identifier(identifier: str) -> tuple[str, str]:
    if identifier.startswith("GV.") or identifier.startswith("ID."):
        return ("AV_MONETARY;EP_FREQ;HF_AWARENESS", "AssetValue")
    if identifier.startswith("PR.") or identifier.startswith("DE."):
        return ("EP_EXPLOIT;DR;FPR;AUROC;PR_AUC", "EventPotential")
    if identifier.startswith("RS.") or identifier.startswith("RC."):
        return ("IL_CIA;RRL_PORTFOLIO;TCO_3Y", "ImpactLevel")
    if identifier.startswith("A.5"):
        return ("HF_AWARENESS;TCO_3Y", "HumanFactorMetric")
    if identifier.startswith("A.6"):
        return ("HF_AWARENESS;DR", "HumanFactorMetric")
    if identifier.startswith("A.7"):
        return ("IL_CIA;RRL_PORTFOLIO", "ImpactLevel")
    if identifier.startswith("A.8"):
        return ("EP_EXPLOIT;DR;FPR;AUROC", "EventPotential")
    return ("AV_MONETARY;EP_FREQ;IL_CIA", "Risk")


def build_mapping_rows(
    iso_rows: list[dict[str, str]],
    nist_rows: list[dict[str, str]],
) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    iso_to_nist = build_iso_to_nist_lookup(nist_rows)
    for position, row in enumerate(iso_rows, start=1):
        metric_ids, uml_class = metric_bundle_for_identifier(row["id"])
        rows.append(
            {
                "framework": "ISO27001:2022",
                "id": row["id"],
                "title": row["title"],
                "metric_ids": metric_ids,
                "uml_class": uml_class,
                "gqm_ref": f"GQM-ISO-{position:03d}",
                "linked_ids": ";".join(iso_to_nist.get(row["id"], [])),
            }
        )
    for position, row in enumerate(nist_rows, start=1):
        metric_ids, uml_class = metric_bundle_for_identifier(row["id"])
        rows.append(
            {
                "framework": "NIST-CSF-2.0",
                "id": row["id"],
                "title": row["title"],
                "metric_ids": metric_ids,
                "uml_class": uml_class,
                "gqm_ref": f"GQM-CSF-{position:03d}",
                "linked_ids": row["linked_iso_ids"],
            }
        )
    return rows


def build_uml_xmi() -> str:
    root = ET.Element("xmi:XMI", attrib={"xmlns:xmi": "http://www.omg.org/XMI"})
    model = ET.SubElement(root, "uml:Model", attrib={"name": "AIEnabledRiskManagement", "xmlns:uml": "http://www.omg.org/spec/UML/20161101"})
    package = ET.SubElement(model, "packagedElement", attrib={"xmi:type": "uml:Package", "name": "RiskKnowledgeBase"})
    for index, class_name in enumerate(UML_CLASSES, start=1):
        ET.SubElement(
            package,
            "packagedElement",
            attrib={"xmi:type": "uml:Class", "xmi:id": f"class_{index}", "name": class_name},
        )
    return ET.tostring(root, encoding="unicode")


def build_control_catalog() -> list[dict[str, str | float]]:
    specs = [
        ("CTRL-01", "Identity hardening", 28000, 5200, 1600, 0.18, 0.030, -0.004, 0.02, "A.5.15", "PR.AA-01"),
        ("CTRL-02", "MFA expansion", 32000, 6100, 1200, 0.22, 0.040, -0.006, 0.04, "A.5.17", "PR.AA-03"),
        ("CTRL-03", "Security awareness program", 15000, 4200, 900, 0.14, 0.010, 0.000, 0.12, "A.6.3", "PR.AT-01"),
        ("CTRL-04", "Patch orchestration", 26000, 4800, 1800, 0.19, 0.025, -0.003, 0.01, "A.8.8", "PR.PS-02"),
        ("CTRL-05", "Network segmentation", 42000, 6900, 2200, 0.25, 0.018, -0.002, 0.00, "A.8.20", "PR.IR-01"),
        ("CTRL-06", "EDR telemetry uplift", 39000, 7800, 3200, 0.24, 0.052, 0.003, 0.00, "A.8.16", "DE.CM-09"),
        ("CTRL-07", "SIEM tuning", 24000, 5600, 2800, 0.16, 0.035, -0.008, 0.00, "A.8.15", "DE.AE-06"),
        ("CTRL-08", "Vulnerability prioritization", 18000, 3900, 1100, 0.15, 0.012, -0.001, 0.00, "A.5.7", "ID.RA-01"),
        ("CTRL-09", "Backup resilience", 35000, 4500, 1400, 0.17, 0.000, 0.000, 0.00, "A.8.13", "RC.RP-03"),
        ("CTRL-10", "IR tabletop drills", 11000, 2700, 400, 0.11, 0.005, 0.000, 0.03, "A.5.24", "RS.MA-01"),
        ("CTRL-11", "Threat intel enrichment", 16000, 3100, 1400, 0.12, 0.016, 0.001, 0.00, "A.5.7", "ID.RA-03"),
        ("CTRL-12", "Asset inventory cleanup", 14000, 2100, 300, 0.09, 0.003, -0.001, 0.00, "A.5.9", "ID.AM-01"),
        ("CTRL-13", "Supplier assurance uplift", 21000, 3500, 300, 0.10, 0.004, 0.000, 0.01, "A.5.19", "GV.SC-06"),
        ("CTRL-14", "Secure configuration baseline", 27000, 4700, 900, 0.20, 0.020, -0.002, 0.00, "A.8.9", "PR.PS-01"),
    ]
    rows: list[dict[str, str | float]] = []
    for control_id, name, capex, opex, ai_opex, effectiveness, detect_gain, fpr_delta, hf_delta, iso_id, csf_id in specs:
        rows.append(
            {
                "control_id": control_id,
                "name": name,
                "capex": capex,
                "opex": opex,
                "ai_opex": ai_opex,
                "effectiveness": effectiveness,
                "detect_gain": detect_gain,
                "fpr_delta": fpr_delta,
                "hf_delta": hf_delta,
                "linked_iso_id": iso_id,
                "linked_csf_id": csf_id,
                "metric_ids": ";".join(sorted({"RRL_PORTFOLIO", "TCO_3Y", *metric_bundle_for_identifier(iso_id)[0].split(";")})),
            }
        )
    return rows


def build_reference_bundle(root: str | Path, refresh: bool = False) -> ReferencePaths:
    root = Path(root)
    data_dir = root / "data"
    nist_catalog_path = data_dir / "nist_csf_2_0_subcats.csv"
    iso_rows = generate_iso_controls()
    if nist_catalog_path.exists() and not refresh:
        nist_rows = _read_existing_rows(nist_catalog_path)
    else:
        workbook_bytes = fetch_csf_workbook_bytes()
        nist_rows = parse_csf_subcategories(workbook_bytes)
    metric_rows = build_metric_catalog()
    mapping_rows = build_mapping_rows(iso_rows, nist_rows)
    control_rows = build_control_catalog()
    crosswalk_rows = build_crosswalk_rows(iso_rows, nist_rows)
    iso_catalog = _maybe_write_rows(data_dir / "iso27001_2022_annexA.csv", iso_rows, ["id", "title", "domain"], refresh=refresh)
    nist_catalog = _maybe_write_rows(
        nist_catalog_path,
        nist_rows,
        ["id", "title", "function", "category", "informative_references", "linked_iso_ids"],
        refresh=refresh,
    )
    metric_catalog = _maybe_write_rows(data_dir / "metric_catalog.csv", metric_rows, ["metric_id", "description"], refresh=refresh)
    mapping_catalog = _maybe_write_rows(
        data_dir / "mapping_iso_csf_gqm.csv",
        mapping_rows,
        ["framework", "id", "title", "metric_ids", "uml_class", "gqm_ref", "linked_ids"],
        refresh=refresh,
    )
    uml_schema = _maybe_write_text(data_dir / "uml_schema.xmi", build_uml_xmi(), refresh=refresh)
    control_catalog = _maybe_write_rows(
        data_dir / "control_catalog.csv",
        control_rows,
        [
            "control_id",
            "name",
            "capex",
            "opex",
            "ai_opex",
            "effectiveness",
            "detect_gain",
            "fpr_delta",
            "hf_delta",
            "linked_iso_id",
            "linked_csf_id",
            "metric_ids",
        ],
        refresh=refresh,
    )
    crosswalk_catalog = _maybe_write_rows(
        data_dir / "iso_nist_semantic_crosswalk.csv",
        crosswalk_rows,
        ["iso_id", "iso_title", "nist_ids", "crosswalk_source"],
        refresh=refresh,
    )
    _remove_stale_reference_artifacts(data_dir)
    return ReferencePaths(
        iso_catalog=iso_catalog,
        nist_catalog=nist_catalog,
        metric_catalog=metric_catalog,
        mapping_catalog=mapping_catalog,
        uml_schema=uml_schema,
        control_catalog=control_catalog,
        crosswalk_catalog=crosswalk_catalog,
    )


def parse_uml_class_names(xmi_path: str | Path) -> set[str]:
    tree = ET.parse(xmi_path)
    root = tree.getroot()
    class_names: set[str] = set()
    for element in root.iter():
        if element.attrib.get("name") and element.attrib.get("{http://www.omg.org/XMI}type", "").endswith("Class"):
            class_names.add(element.attrib["name"])
        elif element.attrib.get("name") and element.attrib.get("xmi:type", "").endswith("Class"):
            class_names.add(element.attrib["name"])
    return class_names


def flatten_metric_ids(raw_value: str) -> list[str]:
    return [part.strip() for part in raw_value.split(";") if part.strip()]


def iter_unique_ids(rows: Iterable[dict[str, str]]) -> set[str]:
    return {row["id"] for row in rows}


def flatten_linked_ids(raw_value: str) -> list[str]:
    return [part.strip() for part in str(raw_value).split(";") if part.strip()]


def build_iso_to_nist_lookup(nist_rows: list[dict[str, str]]) -> dict[str, list[str]]:
    lookup: dict[str, list[str]] = {}
    for row in nist_rows:
        nist_id = row["id"]
        for iso_id in flatten_linked_ids(row.get("linked_iso_ids", "")):
            lookup.setdefault(iso_id, []).append(nist_id)
    return {iso_id: sorted(set(nist_ids)) for iso_id, nist_ids in lookup.items()}


def build_crosswalk_rows(iso_rows: list[dict[str, str]], nist_rows: list[dict[str, str]]) -> list[dict[str, str]]:
    iso_to_nist = build_iso_to_nist_lookup(nist_rows)
    rows: list[dict[str, str]] = []
    for row in iso_rows:
        rows.append(
            {
                "iso_id": row["id"],
                "iso_title": row["title"],
                "nist_ids": ";".join(iso_to_nist.get(row["id"], [])),
                "crosswalk_source": "NIST-CSF-2.0 informative references",
            }
        )
    return rows
